import openpyxl as opx
import sqlite3

def generar_bd(path_excel: str, path_bd: str):

    archivo_excel = opx.load_workbook(path_excel) 
    # seleccionamos el archivo excel de donde queremos extraer los datos

    # CREAR LA BD Y UN CURSOR PARA INTERACCIONAR CON ELLA

    conexion = sqlite3.connect(path_bd) # generamos una nueva base de datos en la ruta especificada y establacemos la conexion
    cursor_beta = conexion.cursor() # generamos el cursor para interactuar

    #CREAR TABLAS

    for hoja in archivo_excel.sheetnames: # recorremos las hojas del excel
        query_create = f'CREATE TABLE {hoja} (\n' # comenzamos a definir la query para generar las tablas, seleccionamos el nombre de las tablas en 
                                                # funcion de las pestañas del excel (cada pestaña comprende los datos de una tabla)
            
        for col in archivo_excel[hoja][1]:  # generamos los campos de las tablas con los campos definidos en la primera linea de la hoja (A1 -> campo 1, B1 -> campo 2) 
            if col == archivo_excel[hoja][1][0]: 
                query_create += f'{col.value} INTEGER PRIMARY KEY,\n' # continuamos construyendo la query con estos campos, el primero se presupone la key
            else:
                query_create += f'{col.value},\n' # continuamos construyendo la query
        
        query_create += ')' # cerramos la query
        query_create = query_create.replace(',\n)', '\n)') # quitamos la ultima coma que se ha quedado en la query

        cursor_beta.execute(query_create) # ejecutamos la query y queda creada la primera tabla, prosigue a la siguiente hoja del excel/ tabña pues estamos en un bucle
    
    conexion.close() # cerramos conexion
    archivo_excel.close() # cerramos el excel


def aniadir_data(path_excel: str, path_bd: str):

    archivo_excel = opx.load_workbook(path_excel) # abrimos el excel

    conexion = sqlite3.connect(path_bd) # accedemos a la base de datos y establacemos la conexion
    cursor_beta = conexion.cursor() # generamos el cursor para interactuar

    # AÑADIR DATOS

    for tabla in archivo_excel.sheetnames: # recorremos las hojas del excel / tablas
        columnas = [celda.value for celda in archivo_excel[tabla][1]] 
        columnas = tuple(columnas)
        columnas = str(columnas) # definimos la primera parte de la query definida por los campos de la tabla

        for line in archivo_excel[tabla].iter_rows(2): # recorremos las lineas de la hoja donde encontramos los datos a excepcion de la primera que contiene el nombre de los campos
            
            celdas = [celda.value for celda in line]
            celdas = tuple(celdas)
            celdas = str(celdas) # generamos la segunda parte de la query donde encontramos los datos a insertar

            query_insert = f'INSERT INTO {tabla} {columnas} VALUES {celdas};' # generamos la query de insercion incluyendo las partes 
            cursor_beta.execute(query_insert) # ejecutamos la query

    conexion.commit() # comiteamos los updates
    conexion.close() # cerramos la conexion
    archivo_excel.close() # cerramos el excel